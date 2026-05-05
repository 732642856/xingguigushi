#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 创建工作簿
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "ComfyUI提示词表"

# 设置表头
headers = ["镜头编号", "场景", "景别", "正向提示词", "负向提示词", "推荐参数"]
ws.append(headers)

# 设置表头样式
header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

# 所有提示词数据
all_prompts = [
    # 场景1 (8镜)
    ["001", "场景1-废墟开场", "特写", 
     "Close-up of a worn gas mask filling the frame, scratched glass visor covered in dust that cannot be wiped off, post-apocalyptic wasteland lighting, desaturated color palette, gritty texture, cinematic lighting from the side, shallow depth of field, 35mm film look, photorealistic, 8k, highly detailed",
     "cartoon, anime, illustration, clean, bright colors, modern, cheerful, deformed, blurry, low quality",
     "模型: z-image/LTX 2.3\n采样器: dpmpp_2m\n步数: 30\nCFG: 7.5\n分辨率: 1920x1080"],
    
    ["002", "场景1-废墟开场", "特写→近景", 
     "Close-up pulling back to reveal a man's face, 38 years old, unshaven beard like weeds, radiation burn scar on left cheek with dark edges, hollow dead eyes like a switched-off light, post-apocalyptic wasteland background, desaturated yellow-gray tones, cinematic lighting, dramatic shadows, photorealistic, 8k, highly detailed skin texture",
     "cartoon, anime, illustration, clean-shaven, healthy, bright, cheerful, deformed, blurry",
     "模型: z-image/LTX 2.3\n采样器: dpmpp_2m\n步数: 35\nCFG: 7.0\n分辨率: 1920x1080\n注意: 特写转中景的过渡帧"],
    
    ["003", "场景1-废墟开场", "远景", 
     "Wide establishing shot from low angle behind a lone figure standing before collapsed overpass, broken bridge with exposed steel rebar like broken bones, post-apocalyptic city ruins stretching to horizon, desaturated gray-yellow sky, dusty atmosphere, cinematic composition, rule of thirds, epic scale, 35mm film look, photorealistic, 8k",
     "cartoon, anime, illustration, bright sky, clean city, modern buildings, cheerful, deformed",
     "模型: z-image/LTX 2.3\n采样器: dpmpp_2m\n步数: 40\nCFG: 7.0\n分辨率: 1920x1080\n宽银幕构图 2.39:1"],
    
    ["004", "场景1-废墟开场", "特写", 
     "Extreme close-up of a wrist-mounted radiation detector, digital display showing numbers flickering 47-48-47 μSv/h, yellow warning color, blinking indicator light, post-apocalyptic worn device with scratches, cinematic macro shot, shallow depth of field, photorealistic, 8k, highly detailed",
     "cartoon, anime, illustration, clean device, modern UI, bright colors, cheerful, deformed, blurry",
     "模型: z-image/LTX 2.3\n采样器: dpmpp_2m\n步数: 25\nCFG: 8.0\n分辨率: 1920x1080\n微距效果"],
    
    ["005", "场景1-废墟开场", "中景→近景", 
     "Medium shot following a man walking through abandoned vehicles, passing a car with broken windows, skeleton in driver's seat wearing decayed suit and tie, focus shift from foreground to background, post-apocalyptic wasteland, desaturated colors, dusty atmosphere, cinematic lighting, photorealistic, 8k",
     "cartoon, anime, illustration, living people, clean cars, modern city, cheerful, deformed",
     "模型: z-image/LTX 2.3\n采样器: dpmpp_2m\n步数: 35\nCFG: 7.0\n分辨率: 1920x1080\n焦点变换效果"],
    
    ["006", "场景1-废墟开场", "全景", 
     "Wide shot from high angle looking down at a massive mutated crow perched on utility pole, wingspan nearly one meter, featherless skin like rotting leather, glowing red eyes staring coldly, post-apocalyptic wasteland background, desaturated colors, cinematic lighting, dramatic atmosphere, photorealistic, 8k",
     "cartoon, anime, illustration, normal crow with feathers, healthy bird, bright colors, cheerful, deformed",
     "模型: z-image/LTX 2.3\n采样器: dpmpp_2m\n步数: 35\nCFG: 7.5\n分辨率: 1920x1080\n高角度俯拍"],
    
    ["007", "场景1-废墟开场", "远景→大远景", 
     "Wide shot pulling back to extreme wide, lone figure walking through post-apocalyptic wasteland becoming smaller and smaller, ruins stretching to horizon, cockroach crawling over decayed shoes in foreground, desaturated gray-yellow sky, sun like blurred smudge obscured by radiation dust, cinematic composition, epic scale, photorealistic, 8k",
     "cartoon, anime, illustration, clean environment, green nature, bright sky, cheerful, deformed",
     "模型: z-image/LTX 2.3\n采样器: dpmpp_2m\n步数: 40\nCFG: 6.5\n分辨率: 1920x1080\n拉远镜头，强调渺小感"],
    
    # 场景2 (10镜)
    ["008", "场景2-废弃超市", "特写→全景", 
     "Close-up of broken glass door pulling back to reveal interior of abandoned supermarket, half-fallen sign with only one character remaining, camera pushing through shattered glass into building, dark interior with light filtering through high windows, post-apocalyptic decay, cinematic lighting, photorealistic, 8k",
     "cartoon, anime, illustration, clean supermarket, bright lights, stocked shelves, cheerful customers, deformed",
     "模型: z-image/LTX 2.3\n采样器: dpmpp_2m\n步数: 35\nCFG: 7.0\n分辨率: 1920x1080\n推轨进入效果"],
    
    ["009", "场景2-废弃超市", "全景→中景", 
     "Interior of abandoned supermarket with collapsed shelves like bomb damage, scattered products with faded packaging, rat running with object in mouth toward corner, dust particles in air, dim lighting from high windows, post-apocalyptic atmosphere, cinematic composition, photorealistic, 8k",
     "cartoon, anime, illustration, clean store, bright lights, organized shelves, cheerful, deformed",
     "模型: z-image/LTX 2.3\n采样器: dpmpp_2m\n步数: 35\nCFG: 7.0\n分辨率: 1920x1080\n昏暗内部光线"],
    
    ["010", "场景2-废弃超市", "特写", 
     "Extreme close-up of man's nose and face, he sniffs the air catching a scent, eyebrows furrowing slightly, first expression change in years, dim post-apocalyptic lighting, shallow depth of field, cinematic macro shot, photorealistic, 8k, highly detailed skin texture",
     "cartoon, anime, illustration, expressionless, smooth skin, bright lighting, cheerful, deformed",
     "模型: z-image/LTX 2.3\n采样器: dpmpp_2m\n步数: 30\nCFG: 7.5\n分辨率: 1920x1080\n表情特写"],
    
    ["011", "场景2-废弃超市", "特写", 
     "POV shot of hands pushing aside collapsed supermarket shelf, revealing shadowy space behind, blue baby swaddle visible in darkness, gently moving, single spot of color in gray environment, post-apocalyptic atmosphere, cinematic lighting, photorealistic, 8k",
     "cartoon, anime, illustration, bright colors everywhere, cheerful scene, no baby, deformed",
     "模型: z-image/LTX 2.3\n采样器: dpmpp_2m\n步数: 35\nCFG: 7.0\n分辨率: 1920x1080\n蓝色襁褓是唯一亮色"],
    
    ["012", "场景2-废弃超市", "特写→大特写", 
     "Close-up pushing in to extreme close-up of man's eyes, pupils constricting in shock, like seeing oasis after wandering desert for days, tears forming but not falling, dim post-apocalyptic lighting with single light source, cinematic lighting, shallow depth of field, photorealistic, 8k",
     "cartoon, anime, illustration, happy expression, bright eyes, cheerful lighting, deformed",
     "模型: z-image/LTX 2.3\n采样器: dpmpp_2m\n步数: 35\nCFG: 7.5\n分辨率: 1920x1080\n瞳孔收缩震惊效果"],
    
    ["013", "场景2-废弃超市", "大特写", 
     "Overhead shot slowly tilting down, trembling hands lifting corner of swaddle revealing baby's face, pink skin, closed eyes, peaceful breathing, about 7 months old, slightly thin but alive, single shaft of gray light illuminating face like spotlight, cinematic lighting, photorealistic, 8k",
     "cartoon, anime, illustration, sick baby, unhealthy, dark lighting, deformed features, blurry",
     "模型: z-image/LTX 2.3\n采样器: dpmpp_2m\n步数: 35\nCFG: 7.0\n分辨率: 1920x1080\n俯拍+缓慢下摇"],
    
    ["014", "场景2-废弃超市", "近景", 
     "Medium shot from side of man pulling back hand and stepping backward, internal struggle visible on face, looking at exit then at baby, lips moving but no sound, post-apocalyptic supermarket interior, dim lighting, cinematic composition, photorealistic, 8k",
     "cartoon, anime, illustration, happy expression, bright lighting, cheerful, deformed",
     "模型: z-image/LTX 2.3\n采样器: dpmpp_2m\n步数: 30\nCFG: 7.0\n分辨率: 1920x1080\n内心挣扎表情"],
    
    ["015", "场景2-废弃超市", "特写群", 
     "Flashback sequence in black and white with film grain, quick cuts: woman's smiling face → baby's small hand → explosion light → screaming mouth → black screen, vintage film look, shaking frame, burned edges, desaturated, old film texture, cinematic, 8k",
     "cartoon, anime, illustration, color footage, modern look, clean image, cheerful, deformed",
     "模型: z-image/LTX 2.3\n采样器: dpmpp_2m\n步数: 30\nCFG: 7.0\n分辨率: 1920x1080\n黑白闪回+胶片质感"],
    
    ["016", "场景2-废弃超市", "特写", 
     "Close-up between man and baby, baby opening eyes wide with curiosity, clear innocent gaze, small hand reaching out and grabbing man's finger, shallow depth of field with focus on hands, soft lighting, emotional moment, photorealistic, 8k",
     "cartoon, anime, illustration, crying baby, scared expression, dark lighting, deformed hands, blurry",
     "模型: z-image/LTX 2.3\n采样器: dpmpp_2m\n步数: 35\nCFG: 7.0\n分辨率: 1920x1080\n手实焦眼睛虚焦"],
    
    ["017", "场景2-废弃超市", "大特写", 
     "Extreme close-up of man's eyes slowly pushing in, slightly red rimmed, first urge to cry in 7 years, single tear forming but not falling, light from behind illuminating side of face like redemption, cinematic lighting, shallow depth of field, photorealistic, 8k",
     "cartoon, anime, illustration, happy tears, bright lighting, cheerful, deformed",
     "模型: z-image/LTX 2.3\n采样器: dpmpp_2m\n步数: 35\nCFG: 7.5\n分辨率: 1920x1080\n眼眶微红想哭的效果"],
]

# 场景3 (12镜)
scene3_prompts = [
    ["018", "场景3-荒野逃亡", "特写", 
     "Low angle close-up of running feet in military boots, stepping over碎石 glass dead rat, motion blur, dust kicking up, post-apocalyptic wasteland ground, gritty texture, cinematic motion shot, photorealistic, 8k",
     "cartoon, anime, illustration, clean shoes, modern sneakers, bright clean ground, cheerful, deformed, blurry",
     "模型: z-image/LTX 2.3\n采样器: dpmpp_2m\n步数: 25\nCFG: 7.0\n分辨率: 1920x1080\n运动模糊"],
    
    ["019", "场景3-荒野逃亡", "中景", 
     "Side view medium shot following man running on wasteland highway carrying baby, protecting baby's head with hand even while stumbling, abandoned vehicles like steel graves in background, some with skeletons inside, desaturated colors, dust in air, cinematic lighting, photorealistic, 8k",
     "cartoon, anime, illustration, clean highway, modern cars, green nature, cheerful, deformed",
     "模型: z-image/LTX 2.3\n采样器: dpmpp_2m\n步数: 30\nCFG: 7.0\n分辨率: 1920x1080\n跟随拍摄"],
    
    ["020", "场景3-荒野逃亡", "特写", 
     "Close-up of man frequently checking wrist-mounted radiation detector, display showing 58 μSv/h in flashing red, yellow warning, cinematic macro shot, shallow depth of field, urgent atmosphere, photorealistic, 8k",
     "cartoon, anime, illustration, clean device, normal readings, bright colors, cheerful, deformed, blurry",
     "模型: z-image/LTX 2.3\n采样器: dpmpp_2m\n步数: 25\nCFG: 8.0\n分辨率: 1920x1080\n红色警报闪烁"],
    
    ["021", "场景3-荒野逃亡", "近景", 
     "Medium shot behind abandoned truck, man catching breath, taking out canteen and feeding baby water with cap, close-up insert of baby's lips cracked and sucking on cap, man's eyes softening like ice cracking, post-apocalyptic wasteland, cinematic lighting, photorealistic, 8k",
     "cartoon, anime, illustration, clean water bottle, healthy baby, bright lighting, cheerful, deformed",
     "模型: z-image/LTX 2.3\n采样器: dpmpp_2m\n步数: 35\nCFG: 7.0\n分辨率: 1920x1080\n柔和眼神"],
    
    ["022", "场景3-荒野逃亡", "特写", 
     "POV close-up of old photograph in man's hand, family of three, heavily worn edges blurred, photo turning to color showing vivid memory then back to black and white, vintage photo texture, cinematic lighting, emotional, photorealistic, 8k",
     "cartoon, anime, illustration, new photo, bright colors, modern family photo, cheerful, deformed",
     "模型: z-image/LTX 2.3\n采样器: dpmpp_2m\n步数: 30\nCFG: 7.0\n分辨率: 1920x1080\n照片变彩色效果"],
    
    ["023", "场景3-荒野逃亡", "远景", 
     "Wide shot from behind man, two mutated dogs appearing at end of highway in distance, red eyes glowing in gray darkness, post-apocalyptic wasteland atmosphere, desaturated colors, threatening atmosphere, cinematic composition, photorealistic, 8k",
     "cartoon, anime, illustration, normal dogs, cute puppies, bright colors, cheerful, deformed",
     "模型: z-image/LTX 2.3\n采样器: dpmpp_2m\n步数: 35\nCFG: 7.5\n分辨率: 1920x1080\n变异犬红眼睛发光"],
    
    ["024", "场景3-荒野逃亡", "大特写", 
     "Extreme close-up of man's face, tense expression but no fear, he recognizes this sound, lips moving whispering, dim post-apocalyptic lighting, cinematic lighting, shallow depth of field, photorealistic, 8k",
     "cartoon, anime, illustration, terrified expression, bright lighting, cheerful, deformed",
     "模型: z-image/LTX 2.3\n采样器: dpmpp_2m\n步数: 30\nCFG: 7.5\n分辨率: 1920x1080\n紧张但不恐惧的表情"],
    
    ["025", "场景3-荒野逃亡", "全景", 
     "High angle wide shot looking down at man weaving through ruins familiar with terrain, jumping over abandoned car with agile movement, quick cut inserts showing sequence of jumps, fast pace tense like heartbeat, post-apocalyptic wasteland, cinematic composition, photorealistic, 8k",
     "cartoon, anime, illustration, clean park, green grass, bright colors, cheerful, deformed",
     "模型: z-image/LTX 2.3\n采样器: dpmpp_2m\n步数: 30\nCFG: 7.0\n分辨率: 1920x1080\n高角度俯视+快速横移"],
    
    ["026", "场景3-荒野逃亡", "中景", 
     "Side view medium shot following mutated dogs increasing to five, scarred leader fastest closing distance, nearest one lunging at man who dodges sideways and hits dog with backpack, slow motion of backpack tearing cans and canteens scattering, post-apocalyptic wasteland, cinematic lighting, photorealistic, 8k",
     "cartoon, anime, illustration, normal dogs playing, cute animals, bright colors, cheerful, deformed",
     "模型: z-image/LTX 2.3\n采样器: dpmpp_2m\n步数: 35\nCFG: 7.0\n分辨率: 1920x1080\n慢动作背包击中效果"],
    
    ["027", "场景3-荒野逃亡", "近景", 
     "Front view medium shot following baby starting to cry, close-up insert of baby's face showing fear but also dependence on man, man patting baby while running, voice breathless but trying to be gentle, post-apocalyptic wasteland, cinematic lighting, emotional, photorealistic, 8k",
     "cartoon, anime, illustration, happy laughing baby, bright lighting, cheerful, deformed, blurry",
     "模型: z-image/LTX 2.3\n采样器: dpmpp_2m\n步数: 35\nCFG: 7.0\n分辨率: 1920x1080\n哭泣婴儿特写"],
    
    ["028", "场景3-荒野逃亡", "远景", 
     "Wide shot from behind, broken bridge ahead, weak yellow light on other side like star in darkness, man's eyes lighting up with hope, post-apocalyptic wasteland, desaturated colors with single point of warm light, cinematic composition, photorealistic, 8k",
     "cartoon, anime, illustration, modern bridge, bright lights, cityscape, cheerful, deformed",
     "模型: z-image/LTX 2.3\n采样器: dpmpp_2m\n步数: 40\nCFG: 6.5\n分辨率: 1920x1080\n远处希望之光"],
    
    ["029", "场景3-荒野逃亡", "特写群", 
     "Quick cut sequence from multiple angles: man's running face in focus → dog legs blurred → close-up of baby crying → detector alarm flashing → man panting close-up → hand protecting baby's head, sounds interweaving rhythm getting faster like drumbeat, fast editing, intense, cinematic, 8k",
     "cartoon, anime, illustration, slow pace, peaceful scene, bright colors, cheerful, deformed",
     "模型: z-image/LTX 2.3\n采样器: dpmpp_2m\n步数: 25\nCFG: 7.0\n分辨率: 1920x1080\n快速剪辑"],
]

# 场景4 (7镜)
scene4_prompts = [
    ["030", "场景4-断桥飞跃", "特写→全景", 
     "Close-up pulling back to wide shot at bridge edge, broken bridge with three-meter gap in middle, abyss below pitch black with only wind sound like ghost crying, dusk sky turning orange-yellow like blood, post-apocalyptic wasteland, cinematic composition, photorealistic, 8k",
     "cartoon, anime, illustration, intact bridge, bright day, cheerful, deformed",
     "模型: z-image/LTX 2.3\n采样器: dpmpp_2m\n步数: 35\nCFG: 7.0\n分辨率: 1920x1080\n黄昏橙黄色天空"],
    
    ["031", "场景4-断桥飞跃", "近景", 
     "Medium shot front view of man reaching bridge edge panting, close-up insert of radiation detector showing疯狂 72 μSv/h in flashing red, time running out, pale face with sweat, urgent atmosphere, cinematic lighting, photorealistic, 8k",
     "cartoon, anime, illustration, normal readings, calm situation, bright colors, cheerful, deformed",
     "模型: z-image/LTX 2.3\n采样器: dpmpp_2m\n步数: 30\nCFG: 8.0\n分辨率: 1920x1080\n疯狂闪烁红色警报"],
    
    ["032", "场景4-断桥飞跃", "全景", 
     "Wide high angle shot looking down, seven mutated dogs surrounding from all directions, scarred leader slowly approaching growling, camera orbiting 360 degrees around man revealing circle of dogs, man trapped like cornered beast, post-apocalyptic wasteland, dusk lighting, cinematic composition, photorealistic, 8k",
     "cartoon, anime, illustration, normal dogs, friendly animals, bright colors, cheerful, deformed",
     "模型: z-image/LTX 2.3\n采样器: dpmpp_2m\n步数: 40\nCFG: 7.5\n分辨率: 1920x1080\n环绕360度镜头"],
    
    ["033", "场景4-断桥飞跃", "近景", 
     "Side view medium shot of man's face showing despair but determined eyes, looking down at baby whimpering small hand clutching his clothes, man bowing head to kiss baby's forehead, gentle father's tenderness returning after 7 years, post-apocalyptic wasteland, dusk lighting, emotional, cinematic lighting, photorealistic, 8k",
     "cartoon, anime, illustration, happy expression, bright lighting, cheerful, deformed",
     "模型: z-image/LTX 2.3\n采样器: dpmpp_2m\n步数: 35\nCFG: 7.0\n分辨率: 1920x1080\n亲吻额头温情时刻"],
    
    ["034", "场景4-断桥飞跃", "特写", 
     "Close-up front view of man's eyes showing determination, sunset illuminating face like final judgment light, red orange golden hour lighting, dramatic shadows, cinematic lighting, shallow depth of field, photorealistic, 8k",
     "cartoon, anime, illustration, scared eyes, dark lighting, cheerful, deformed",
     "模型: z-image/LTX 2.3\n采样器: dpmpp_2m\n步数: 35\nCFG: 7.5\n分辨率: 1920x1080\n夕阳金色时刻光照"],
    
    ["035", "场景4-断桥飞跃", "全景→中景", 
     "Wide to medium side view shot, slow motion of man taking few steps back then sprinting and leaping off broken bridge, slow motion breakdown: running → jumping → looking back at dogs in air → tightly hugging baby wrapping body around her → continuing to fall, sunset backlighting creating golden silhouette like angel martyr, dust and debris floating in slow motion like snowflakes, cinematic epic shot, photorealistic, 8k",
     "cartoon, anime, illustration, normal speed, bright day, cheerful, deformed",
     "模型: z-image/LTX 2.3\n采样器: dpmpp_2m\n步数: 45\nCFG: 7.0\n分辨率: 1920x1080\n慢动作15秒分解"],
    
    ["036", "场景4-断桥飞跃", "中景→近景", 
     "Medium to close-up shot from low ground angle, man landing and rolling using back to absorb impact protecting baby, arm scraped bleeding staining ground red, immediately getting up to check baby, close-up insert of baby coughing and opening eyes looking at man, man: almost there, man's eyes red difficult but genuine smile, post-apocalyptic wasteland, cinematic lighting, emotional, photorealistic, 8k",
     "cartoon, anime, illustration, happy landing, bright lighting, cheerful, deformed",
     "模型: z-image/LTX 2.3\n采样器: dpmpp_2m\n步数: 35\nCFG: 7.0\n分辨率: 1920x1080\n落地翻滚保护婴儿"],
]

# 场景5 (6镜)
scene5_prompts = [
    ["037", "场景5-避难所入口", "特写", 
     "Close-up of heavy metal shelter door, thick door with red indicator light flashing on and off, man's trembling wounded hand knocking, knocking sound heavy like drum like heartbeat, post-apocalyptic shelter entrance, cinematic lighting, photorealistic, 8k",
     "cartoon, anime, illustration, modern door, bright colors, cheerful, deformed",
     "模型: z-image/LTX 2.3\n采样器: dpmpp_2m\n步数: 30\nCFG: 7.5\n分辨率: 1920x1080\n红色指示灯闪烁"],
    
    ["038", "场景5-避难所入口", "全景", 
     "Wide shot from behind man, door slowly opening, backlighting five people standing inside surrounded by warm yellow light, leading white-haired elderly woman with kind face wearing white coat modified robe, behind her middle-aged man guard young woman nurse two children, all clean eyes clean without apocalyptic numbness and fear, strong backlighting figures as silhouettes only outlines but conveying safety and warmth, first appearance of warm tones orange yellow like home color, cinematic composition, photorealistic, 8k",
     "cartoon, anime, illustration, dark interior, cold lighting, fearful people, deformed",
     "模型: z-image/LTX 2.3\n采样器: dpmpp_2m\n步数: 40\nCFG: 6.5\n分辨率: 1920x1080\n强烈逆光剪影效果"],
    
    ["039", "场景5-避难所入口", "近景", 
     "Medium shot from side, elderly woman extending hand to receive baby, man hesitating moment looking down at baby baby also looking at him, then man gently handing baby over, close-up insert of woman receiving baby professional examination finger touching pulse showing relieved smile, father's letting go but also trust, warm lighting, emotional, cinematic lighting, photorealistic, 8k",
     "cartoon, anime, illustration, sad separation, dark lighting, cheerful, deformed",
     "模型: z-image/LTX 2.3\n采样器: dpmpp_2m\n步数: 35\nCFG: 7.0\n分辨率: 1920x1080\n温情交接"],
    
    ["040", "场景5-避难所入口", "大特写", 
     "Extreme close-up slowly pushing in, baby looking at man suddenly smiling, crisp laughter like bells like light, laughter first time in entire film like music like redemption, warm light shining on baby's face gilding skin with golden color, emotional moment, cinematic lighting, shallow depth of field, photorealistic, 8k",
     "cartoon, anime, illustration, crying baby, dark lighting, cheerful, deformed, blurry",
     "模型: z-image/LTX 2.3\n采样器: dpmpp_2m\n步数: 35\nCFG: 7.0\n分辨率: 1920x1080\n婴儿笑声天使般"],
    
    ["041", "场景5-避难所入口", "大特写", 
     "Extreme close-up of man's face, eye sockets moist face showing first genuine relieved smile in 7 years, man: little light... naming redemption complete new life beginning, warm lighting, emotional climax, cinematic lighting, shallow depth of field, photorealistic, 8k",
     "cartoon, anime, illustration, sad crying, dark lighting, cheerful, deformed",
     "模型: z-image/LTX 2.3\n采样器: dpmpp_2m\n步数: 35\nCFG: 7.0\n分辨率: 1920x1080\n真心释然笑容"],
    
    ["042", "场景5-避难所入口", "全景→大远景", 
     "Wide to extreme wide shot from side, man surrounded by five shelter residents, someone handing water someone supporting him someone looking at him with concern, camera slowly pulling back reveal foreground crowd warm light background metal door, extreme background outside window post-apocalyptic wasteland gray yellow death despair, subtitles: in darkness there is always light dedicated to all who guard hope in despair, hope inside window despair outside window man walked from outside to inside, fade to black, cinematic epic ending, photorealistic, 8k",
     "cartoon, anime, illustration, all despair, no hope, cheerful, deformed",
     "模型: z-image/LTX 2.3\n采样器: dpmpp_2m\n步数: 45\nCFG: 6.5\n分辨率: 1920x1080\n最终拉远镜头对比"],
]

# 合并所有数据
all_prompts = all_prompts + scene3_prompts + scene4_prompts + scene5_prompts

# 添加所有数据
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for row_data in all_prompts:
    ws.append(row_data)

# 设置所有单元格样式
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.border = thin_border

# 设置列宽
column_widths = [10, 15, 10, 60, 40, 30]
for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# 设置行高
ws.row_dimensions[1].height = 30
for row in range(2, ws.max_row + 1):
    ws.row_dimensions[row].height = 80

# 保存文件
output_path = "/Users/tony/WorkBuddy/20260411141627/剧本项目/末世之光_20260411/ComfyUI提示词表.xlsx"
wb.save(output_path)
print(f"ComfyUI提示词Excel文件已保存至: {output_path}")
print(f"总镜头数: {len(all_prompts)}")
