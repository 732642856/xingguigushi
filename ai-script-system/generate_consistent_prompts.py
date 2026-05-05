#!/usr/bin/env python3
"""
《末世之光》ComfyUI 一致性提示词生成器 v2.0
确保角色在不同镜头中保持外貌一致性
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============ 角色一致性特征定义 ============
CHEN_BASE = """38-year-old Asian male survivor, thin gaunt face with prominent cheekbones, unshaven beard like wild weeds, messy dark hair with white strands, distinctive 3cm×2cm radiation burn scar on left cheek (purplish-red rough patch), wearing cracked wire-rimmed glasses with spiderweb crack on left lens, weathered khaki protective coat with multiple patches, wrist-mounted radiation detector with cracked screen, weathered backpack with repaired straps, rusted dagger at waist"""

CHEN_FACE = """hollow exhausted eyes, weathered skin with visible aging from 7 years wasteland survival, prominent cheekbones, unshaven beard"""

CHEN_CLOTHING = """tattered khaki protective suit coat with patches, worn military boots, cracked wire-rimmed glasses"""

CHEN_EQUIPMENT = """cracked radiation detector on left wrist, patched backpack, rusted dagger at waist"""

BABY_BASE = """7-month-old infant with fair pinkish skin slightly pale from malnutrition, sparse golden blonde fine hair, large bright eyes with light brown irises reflecting light, wrapped in distinctive sky-blue cotton swaddle with faded star and moon pattern, soft chubby cheeks with baby fat, delicate small hands with tiny fingernails"""

DOG_BASE = """mutated dog 30% larger than normal, shoulder height 75-85cm, hairless body, purplish-black skin covered in festering scabs and ulcers, red glowing eyes without whites, abnormally developed teeth 3-5cm exposed yellow-black with bloody foam, claw-like thickened keratinized nails 5cm curved like sickle blades, short rat-like tail 15cm"""

DOG_LEADER = """largest 85cm shoulder height, blind right eye with cataract, massive scar from eyebrow to chin on right face"""

# ============ 镜头定义 ============
SHOTS = [
    # 场景1
    ("001", "场景1-废墟开场", "远景", None, "Wide establishing shot of post-apocalyptic wasteland cityscape at dusk, abandoned skyscrapers with broken windows silhouetted against gray-orange sky, collapsed highway overpasses, scattered abandoned vehicles covered in dust and rust, dead trees with bare branches, toxic fog hovering over ground, desaturated color palette of grays browns and muted oranges, cinematic composition, photorealistic, highly detailed, 8k quality"),
    
    ("002", "场景1-废墟开场", "特写→近景", "阿尘", f"Close-up pulling back to medium shot of {CHEN_BASE}, {CHEN_FACE} showing expressionless hollow dead eyes like switched-off lights, standing amidst rubble of collapsed concrete building, dust particles floating in dim light, post-apocalyptic wasteland atmosphere, shallow depth of field, cinematic lighting with hard shadows, photorealistic, highly detailed, 8k"),
    
    ("003", "场景1-废墟开场", "全景", "阿尘", f"Wide shot of {CHEN_BASE}, {CHEN_CLOTHING}, walking alone through vast ruined cityscape, tiny figure against massive collapsed buildings, scattered debris and abandoned vehicles, gray overcast sky, dust kicked up with each step, atmospheric haze, sense of isolation and scale, desaturated post-apocalyptic color grading, cinematic wide angle, photorealistic, highly detailed, 8k"),
    
    ("004", "场景1-废墟开场", "特写", "阿尘装备", f"Extreme close-up macro shot of {CHEN_BASE} left wrist, focusing on cracked-screen radiation detector displaying '45 μSv/h' in yellow numbers, screen has spiderweb crack pattern, worn weathered device casing, veins visible on thin weathered hand, shallow depth of field, background blurred showing tattered coat sleeve, cinematic macro photography, photorealistic, highly detailed, 8k"),
    
    ("005", "场景1-废墟开场", "中景→近景", "阿尘", f"Medium shot following {CHEN_BASE}, {CHEN_CLOTHING}, walking past abandoned vehicles with shattered windows, focus shifts to skeleton in driver's seat wearing decayed suit and tie, {CHEN_FACE} glances at skeleton with numb indifferent expression, {CHEN_EQUIPMENT} visible, post-apocalyptic street with scattered debris, cinematic rack focus, desaturated colors, photorealistic, highly detailed, 8k"),
    
    ("006", "场景1-废墟开场", "特写", "变异乌鸦", "Close-up of mutated crow perched on bent streetlight, 1.5m wingspan, feathers partially fallen revealing raw pink skin underneath, one cloudy white blind eye, beak covered in greenish growths, tilted head observing wasteland below, gray toxic sky background, symbolic watcher of the apocalypse, shallow depth of field, cinematic lighting, photorealistic creature design, highly detailed, 8k"),
    
    ("007", "场景1-废墟开场", "近景", "阿尘", f"Medium shot from behind {CHEN_BASE}, {CHEN_CLOTHING}, {CHEN_EQUIPMENT}, standing at entrance looking toward abandoned supermarket, {CHEN_FACE} showing cautious alert expression, left hand instinctively touching dagger handle, overgrown parking lot with wrecked shopping carts, broken store sign hanging by one bolt, evening light casting long shadows, post-apocalyptic atmosphere, cinematic composition, photorealistic, highly detailed, 8k"),
    
    # 场景2
    ("008", "场景2-废弃超市", "全景→中景", "阿尘", f"Wide to medium shot of {CHEN_BASE}, {CHEN_CLOTHING}, {CHEN_EQUIPMENT}, carefully entering abandoned supermarket through shattered glass doors, interior dark with only light from broken skylight above, overturned shelves and scattered merchandise, {CHEN_FACE} alert and scanning surroundings, dust particles in light beams, post-apocalyptic interior, cinematic lighting, photorealistic, highly detailed, 8k"),
    
    ("009", "场景2-废弃超市", "中景", "阿尘", f"Medium shot of {CHEN_BASE}, {CHEN_CLOTHING}, walking through supermarket aisle, shelves mostly empty with some remaining cans and packages covered in dust, {CHEN_FACE} focused and cautious, {CHEN_EQUIPMENT} visible, broken shopping carts and debris on floor, dim lighting from emergency exit signs, post-apocalyptic atmosphere, cinematic composition, photorealistic, highly detailed, 8k"),
    
    ("010", "场景2-废弃超市", "特写", "阿尘", f"Extreme close-up of {CHEN_BASE} nose and face, {CHEN_FACE}, he sniffs the air catching a scent, eyebrows furrowing slightly, first expression change in years, {CHEN_CLOTHING} collar visible, dim post-apocalyptic lighting, shallow depth of field, cinematic macro shot, photorealistic skin texture, highly detailed, 8k"),
    
    ("011", "场景2-废弃超市", "特写", None, "POV shot of hands pushing aside collapsed supermarket shelf, revealing shadowy space behind, blue baby swaddle visible in darkness, gently moving, single spot of color in gray environment, post-apocalyptic supermarket interior, dramatic lighting contrast, cinematic POV composition, photorealistic, highly detailed, 8k"),
    
    ("012", "场景2-废弃超市", "特写→大特写", "阿尘", f"Close-up pushing in to extreme close-up of {CHEN_BASE} eyes, {CHEN_FACE} with cracked glasses, pupils constricting in shock, like seeing oasis after wandering desert for days, tears forming but not falling, dim post-apocalyptic lighting with single shaft of light, shallow depth of field, cinematic eye close-up, photorealistic, highly detailed, 8k"),
    
    ("013", "场景2-废弃超市", "大特写", "小光", f"Overhead shot slowly tilting down, trembling weathered hands of {CHEN_BASE} lifting corner of swaddle revealing {BABY_BASE}, eyes closed, peaceful breathing, single shaft of light illuminating the blue swaddle, post-apocalyptic supermarket interior, dramatic lighting, cinematic overhead angle, photorealistic, highly detailed, 8k"),
    
    ("014", "场景2-废弃超市", "近景", "阿尘+小光", f"Medium shot from side of {CHEN_BASE}, {CHEN_CLOTHING}, {CHEN_EQUIPMENT}, pulling back hand and stepping backward, internal struggle visible on {CHEN_FACE}, looking at exit then at {BABY_BASE}, lips moving but no sound, post-apocalyptic supermarket interior, dim lighting, cinematic two-shot composition, photorealistic, highly detailed, 8k"),
    
    ("015", "场景2-废弃超市", "特写群", "阿尘回忆", f"Flashback sequence in black and white with film grain, quick cuts: woman's smiling face → baby's small hand → explosion light → screaming mouth → black screen, vintage film look, shaking frame, burned edges, {CHEN_BASE} memories, cinematic flashback style, photorealistic, highly detailed, 8k"),
    
    ("016", "场景2-废弃超市", "特写", "阿尘+小光", f"Close-up between {CHEN_BASE}, {CHEN_FACE} with cracked glasses, and {BABY_BASE}, baby opening eyes wide with curiosity, clear innocent gaze, small hand reaching out and grabbing man's weathered finger, shallow depth of field with focus on hands, soft lighting, cinematic two-shot close-up, photorealistic skin textures, highly detailed, 8k"),
    
    ("017", "场景2-废弃超市", "大特写", "阿尘", f"Extreme close-up of {CHEN_BASE} eyes slowly pushing in, {CHEN_FACE} with cracked glasses, slightly red rimmed eyes, first urge to cry in 7 years, single tear forming but not falling, light from behind illuminating side of face like redemption, cinematic eye extreme close-up, photorealistic, highly detailed, 8k"),
    
    ("018", "场景2-废弃超市", "中景", "阿尘+小光", f"Medium shot of {CHEN_BASE}, {CHEN_CLOTHING}, {CHEN_EQUIPMENT}, carefully lifting and holding {BABY_BASE} in arms, protective posture, {CHEN_FACE} showing mixture of fear and determination, post-apocalyptic supermarket interior with shaft of light from above, cinematic medium shot, photorealistic, highly detailed, 8k"),
    
    # 场景3
    ("019", "场景3-荒野逃亡", "中景", "阿尘+小光", f"Side view medium shot following {CHEN_BASE}, {CHEN_CLOTHING}, {CHEN_EQUIPMENT}, running on wasteland highway carrying {BABY_BASE} protectively, {CHEN_FACE} focused and alert, protecting baby's head with hand even while stumbling, abandoned vehicles like steel graves in background, some with skeletons visible, desaturated post-apocalyptic colors, cinematic tracking shot, photorealistic, highly detailed, 8k"),
    
    ("020", "场景3-荒野逃亡", "特写", "阿尘装备", f"Close-up of {CHEN_BASE} hand checking wrist-mounted radiation detector, {CHEN_EQUIPMENT}, display showing '58 μSv/h' in flashing red, yellow warning, {CHEN_FACE} concerned expression visible in background, cinematic macro shot, shallow depth of field, urgent atmosphere, photorealistic, highly detailed, 8k"),
    
    ("021", "场景3-荒野逃亡", "近景", "阿尘+小光", f"Medium shot behind abandoned truck, {CHEN_BASE}, {CHEN_CLOTHING}, catching breath, taking out canteen and feeding {BABY_BASE} water with cap, {CHEN_FACE} softening like ice cracking, baby's small hands reaching up, post-apocalyptic highway with wrecked vehicles, warm moment in harsh environment, cinematic medium shot, photorealistic, highly detailed, 8k"),
    
    ("022", "场景3-荒野逃亡", "特写", "阿尘", f"POV close-up of old photograph in {CHEN_BASE} weathered hand, {CHEN_EQUIPMENT} visible on wrist, family of three photo heavily worn edges blurred, photo turning to color showing vivid memory then back to black and white, {CHEN_FACE} reflecting in photo surface, vintage photo texture, cinematic lighting, photorealistic, highly detailed, 8k"),
    
    ("023", "场景3-荒野逃亡", "远景", "变异犬", f"Wide shot from behind {CHEN_BASE}, {CHEN_CLOTHING}, two {DOG_BASE} appearing at end of highway in distance, red eyes glowing in gray darkness, post-apocalyptic wasteland atmosphere, desaturated colors, threatening atmosphere building, cinematic wide shot with depth, photorealistic creature design, highly detailed, 8k"),
    
    ("024", "场景3-荒野逃亡", "中景", "阿尘+变异犬", f"Medium shot of {CHEN_BASE}, {CHEN_CLOTHING}, {CHEN_EQUIPMENT}, running while looking back over shoulder, {CHEN_FACE} showing fear and determination, {DOG_BASE} visible in background gaining ground, dust kicked up by running, post-apocalyptic highway, cinematic action shot with motion blur, photorealistic, highly detailed, 8k"),
    
    ("025", "场景3-荒野逃亡", "特写", "变异犬", f"Close-up of {DOG_BASE} running, red glowing eyes fixed on prey, purplish-black skin with festering wounds visible, saliva dripping from exposed teeth, motion blur showing speed, post-apocalyptic wasteland background, terrifying creature design, cinematic close-up action shot, photorealistic, highly detailed, 8k"),
    
    ("026", "场景3-荒野逃亡", "近景", "阿尘+小光", f"Medium shot of {CHEN_BASE}, {CHEN_CLOTHING}, {CHEN_EQUIPMENT}, holding {BABY_BASE} close while running, {CHEN_FACE} determined and protective, baby wrapped in blue swaddle calm and quiet, abandoned vehicles blurring past, dust and debris, cinematic tracking medium shot, photorealistic, highly detailed, 8k"),
    
    ("027", "场景3-荒野逃亡", "特写", "阿尘+小光", f"Close-up of {BABY_BASE} in {CHEN_BASE} arms, {CHEN_CLOTHING} visible, baby's small hand gripping man's coat fabric tightly, large innocent eyes looking up at man's {CHEN_FACE} with cracked glasses, expression of trust, shallow depth of field, cinematic intimate close-up, photorealistic, highly detailed, 8k"),
    
    ("028", "场景3-荒野逃亡", "全景", "阿尘+变异犬群", f"Wide shot of {CHEN_BASE}, {CHEN_CLOTHING}, running toward collapsed bridge in distance, pack of 5 {DOG_BASE} in pursuit forming fan shape, red glowing eyes visible even from distance, post-apocalyptic wasteland with dead trees and toxic pools, dramatic sunset lighting, cinematic wide composition showing scale of chase, photorealistic, highly detailed, 8k"),
    
    ("029", "场景3-荒野逃亡", "特写", "阿尘", f"Close-up of {CHEN_BASE} legs and feet in worn military boots running on cracked asphalt, {CHEN_CLOTHING} lower half visible, {DOG_BASE} shadow visible on ground behind, dust and small stones kicked up, motion blur on background, cinematic low angle close-up, photorealistic, highly detailed, 8k"),
    
    # 场景4
    ("030", "场景4-断桥飞跃", "全景", "阿尘+断桥", f"Wide shot of collapsed highway bridge, massive gap with twisted rebar and broken concrete, {CHEN_BASE}, {CHEN_CLOTHING}, standing at edge looking across, {CHEN_EQUIPMENT} visible, {CHEN_FACE} calculating distance, raging toxic river visible far below, post-apocalyptic wasteland stretching to horizon, dramatic cloudy sky, cinematic wide establishing shot, photorealistic, highly detailed, 8k"),
    
    ("031", "场景4-断桥飞跃", "中景", "阿尘+变异犬群", f"Medium shot of {CHEN_BASE}, {CHEN_CLOTHING}, {CHEN_EQUIPMENT}, backing toward bridge edge while {DOG_BASE} including leader with {DOG_LEADER} slowly approach from front, {CHEN_FACE} showing determination mixed with fear, {BABY_BASE} held protectively, pack of 7 dogs forming semicircle, cinematic tense medium shot, photorealistic creature and character design, highly detailed, 8k"),
    
    ("032", "场景4-断桥飞跃", "特写", "变异犬首领", f"Close-up of alpha {DOG_BASE} with {DOG_LEADER}, one blind white eye and massive scar, other red eye glowing with intelligence, snarling with exposed teeth and bloody foam, leader of the pack, most terrifying expression, post-apocalyptic bridge background, cinematic villain close-up, photorealistic creature design, highly detailed, 8k"),
    
    ("033", "场景4-断桥飞跃", "大特写", "阿尘+小光", f"Extreme close-up of {CHEN_BASE} and {BABY_BASE} faces close together, {CHEN_FACE} with cracked glasses showing determination and love, baby's large innocent eyes looking back, moment of connection before leap, shallow depth of field, soft dramatic lighting, cinematic emotional extreme close-up, photorealistic skin textures, highly detailed, 8k"),
    
    ("034", "场景4-断桥飞跃", "全景", "飞跃", f"Wide dramatic shot of {CHEN_BASE}, {CHEN_CLOTHING}, {CHEN_EQUIPMENT}, leaping across gap in collapsed bridge while holding {BABY_BASE} protectively, body suspended in mid-air against dramatic cloudy sky, {DOG_BASE} pack visible at edge behind, slow motion feeling, post-apocalyptic wasteland far below, cinematic wide action shot, photorealistic, highly detailed, 8k"),
    
    ("035", "场景4-断桥飞跃", "中景", "落地", f"Medium shot of {CHEN_BASE}, {CHEN_CLOTHING}, landing hard on other side of bridge, rolling to protect {BABY_BASE} in arms, {CHEN_FACE} showing pain but relief, {CHEN_EQUIPMENT} visible, dust cloud from impact, {DOG_BASE} pack howling from other side unable to follow, cinematic action medium shot, photorealistic, highly detailed, 8k"),
    
    ("036", "场景4-断桥飞跃", "特写", "阿尘", f"Close-up of {CHEN_BASE} lying on ground catching breath, {CHEN_FACE} with cracked glasses showing exhaustion and small smile of victory, {CHEN_CLOTHING} dusty and disheveled, {CHEN_EQUIPMENT} visible, {BABY_BASE} safe in arms, shallow depth of field, cinematic character moment, photorealistic, highly detailed, 8k"),
    
    # 场景5
    ("037", "场景5-避难所入口", "远景", "避难所", f"Wide shot of {CHEN_BASE}, {CHEN_CLOTHING}, {CHEN_EQUIPMENT}, walking toward massive metal door in hillside, {BABY_BASE} in arms, door has warm yellow light leaking from edges, sign of human habitation, {CHEN_FACE} showing hope and exhaustion, post-apocalyptic wasteland transitioning to sanctuary, cinematic wide shot with contrast lighting, photorealistic, highly detailed, 8k"),
    
    ("038", "场景5-避难所入口", "中景", "阿尘", f"Medium shot of {CHEN_BASE}, {CHEN_CLOTHING}, {CHEN_EQUIPMENT}, standing before massive metal door, {CHEN_FACE} with cracked glasses looking up at door with mixture of hope and fear, {BABY_BASE} in arms, weathered hand reaching toward door mechanism, dramatic lighting from door edges, cinematic medium shot, photorealistic, highly detailed, 8k"),
    
    ("039", "场景5-避难所入口", "特写", "开门", f"Close-up of massive metal door slowly opening, bright warm golden light spilling out, silhouettes of people visible inside, {CHEN_BASE} {CHEN_CLOTHING} and {BABY_BASE} visible in foreground as dark shapes against light, dramatic contrast between dark wasteland and warm sanctuary, cinematic lighting contrast shot, photorealistic, highly detailed, 8k"),
    
    ("040", "场景5-避难所入口", "近景", "阿尘+避难所居民", f"Medium shot of {CHEN_BASE}, {CHEN_CLOTHING}, {CHEN_EQUIPMENT}, stepping through doorway into warm light, {CHEN_FACE} expression transforming from exhaustion to wonder, {BABY_BASE} in arms, five clean healthy people visible inside welcoming, warm yellow lighting contrasting with gray exterior, cinematic medium shot, photorealistic, highly detailed, 8k"),
    
    ("041", "场景5-避难所入口", "特写群", "交接", f"Series of close-ups: elderly woman's weathered but kind hands reaching for {BABY_BASE}, {CHEN_BASE} {CHEN_FACE} with cracked glasses reluctant but knowing it's right, hands exchanging the blue swaddle, baby's small hand briefly gripping man's finger before letting go, cinematic emotional close-up sequence, photorealistic skin textures, highly detailed, 8k"),
    
    ("042", "场景5-避难所入口", "大特写", "结局", f"Extreme close-up of {CHEN_BASE} {CHEN_FACE} with cracked glasses, first genuine smile in 7 years, eyes glistening with tears of relief, {CHEN_CLOTHING} visible, warm golden light illuminating face from sanctuary behind, {BABY_BASE} visible in background being held by elderly woman, redemption and hope, cinematic emotional climax, photorealistic, highly detailed, 8k")
]

# ============ 负面提示词 ============
NEGATIVE_PROMPT = "cartoon, anime, illustration, 3d render, cgi, deformed, blurry, low quality, duplicate, watermark, signature, text, cropped, worst quality, low resolution, disfigured, bad anatomy, extra limbs, poorly drawn face, mutation, mutated, extra fingers, fused fingers, too many fingers, long neck, cross-eyed, mutated hands, polar lowres, bad face"

# ============ 推荐参数 ============
PARAMETERS = "Model: z-image/LTX 2.3 | Steps: 30 | CFG: 7.5 | Sampler: DPM++ 2M Karras | Resolution: 1216x704 (16:9 cinematic) | Clip Skip: 2"

def create_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "ComfyUI一致性提示词"
    
    # 标题
    ws['A1'] = "《末世之光》ComfyUI 一致性提示词表 v2.0"
    ws.merge_cells('A1:F1')
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # 说明
    ws['A2'] = "本表确保角色在不同镜头中保持外貌一致性 - 每个角色的核心特征在每个相关镜头中都有完整描述"
    ws.merge_cells('A2:F2')
    ws['A2'].font = Font(size=10, italic=True, color="666666")
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 表头
    headers = ['镜头编号', '场景', '景别', '正向提示词', '负向提示词', '推荐参数']
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 填充数据
    row = 4
    for shot in SHOTS:
        shot_num, scene, shot_type, subject, prompt = shot
        
        ws.cell(row=row, column=1, value=shot_num)
        ws.cell(row=row, column=2, value=scene)
        ws.cell(row=row, column=3, value=shot_type)
        ws.cell(row=row, column=4, value=prompt)
        ws.cell(row=row, column=5, value=NEGATIVE_PROMPT)
        ws.cell(row=row, column=6, value=PARAMETERS)
        
        # 设置样式
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            
            # 根据场景设置背景色
            if "场景1" in scene:
                cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
            elif "场景2" in scene:
                cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            elif "场景3" in scene:
                cell.fill = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")
            elif "场景4" in scene:
                cell.fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
            elif "场景5" in scene:
                cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        
        # 设置行高
        ws.row_dimensions[row].height = 120
        row += 1
    
    # 设置列宽
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 80
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 35
    
    # 添加一致性说明工作表
    ws2 = wb.create_sheet("角色一致性说明")
    
    consistency_info = [
        ["角色一致性特征说明", ""],
        ["", ""],
        ["阿尘 (主角)", "每个包含阿尘的镜头都包含以下完整描述:"],
        ["基础外貌", "38-year-old Asian male, thin gaunt face, unshaven beard, messy dark hair with white strands"],
        ["标志性伤疤", "3cm×2cm radiation burn scar on left cheek (purplish-red rough patch)"],
        ["标志性装备", "cracked wire-rimmed glasses with spiderweb crack on left lens, wrist-mounted radiation detector with cracked screen"],
        ["服装", "weathered khaki protective coat with multiple patches, worn military boots"],
        ["武器/工具", "weathered backpack with repaired straps, rusted dagger at waist"],
        ["", ""],
        ["小光 (婴儿)", "每个包含婴儿的镜头都包含以下完整描述:"],
        ["基础外貌", "7-month-old infant with fair pinkish skin slightly pale from malnutrition, sparse golden blonde fine hair"],
        ["标志性特征", "large bright eyes with light brown irises reflecting light like stars"],
        ["服装", "wrapped in distinctive sky-blue cotton swaddle with faded star and moon pattern"],
        ["身体特征", "soft chubby cheeks with baby fat, delicate small hands with tiny fingernails"],
        ["", ""],
        ["变异犬 (反派)", "每个包含变异犬的镜头都包含以下完整描述:"],
        ["基础外貌", "mutated dog 30% larger than normal, shoulder height 75-85cm, hairless body, purplish-black skin covered in festering scabs"],
        ["标志性特征", "red glowing eyes without whites, abnormally developed teeth 3-5cm exposed yellow-black with bloody foam"],
        ["特殊标记", "claw-like thickened keratinized nails 5cm curved like sickle blades"],
        ["首领特征", "leader: largest 85cm, blind right eye with cataract, massive scar from eyebrow to chin on right face"],
        ["", ""],
        ["使用说明", ""],
        ["1", "所有包含角色的镜头都使用了完全一致的描述模板"],
        ["2", "确保AI生成的角色在不同镜头中保持外貌一致性"],
        ["3", "建议在ComfyUI中使用相同的Seed值或IP-Adapter来进一步增强一致性"],
        ["4", "如需调整角色外貌，只需修改本表上方的CHARACTER_PROFILES定义，所有相关镜头会自动更新"]
    ]
    
    for r_idx, row_data in enumerate(consistency_info, 1):
        for c_idx, value in enumerate(row_data, 1):
            ws2.cell(row=r_idx, column=c_idx, value=value)
    
    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['B'].width = 100
    
    # 保存
    output_path = "/Users/tony/WorkBuddy/20260411141627/剧本项目/末世之光_20260411/ComfyUI提示词表_一致性版.xlsx"
    wb.save(output_path)
    print(f"✅ 已生成一致性提示词表: {output_path}")
    print(f"   共 {len(SHOTS)} 个镜头")
    print(f"   包含阿尘的镜头: {sum(1 for s in SHOTS if s[3] == '阿尘' or s[3] == '阿尘+小光' or s[3] == '阿尘装备' or s[3] == '阿尘回忆' or s[3] == '阿尘+变异犬' or s[3] == '阿尘+变异犬群' or s[3] == '阿尘+断桥' or s[3] == '阿尘+避难所居民')}")
    print(f"   包含婴儿的镜头: {sum(1 for s in SHOTS if s[3] == '小光' or s[3] == '阿尘+小光')}")
    print(f"   包含变异犬的镜头: {sum(1 for s in SHOTS if s[3] == '变异犬' or s[3] == '阿尘+变异犬' or s[3] == '阿尘+变异犬群' or s[3] == '变异犬首领')}")

if __name__ == "__main__":
    create_excel()
