#!/usr/bin/env python3
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "ComfyUI提示词"

# 设置列宽
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 80
ws.column_dimensions['E'].width = 50
ws.column_dimensions['F'].width = 25

# 标题
ws['A1'] = '《深空信号》ComfyUI提示词表'
ws.merge_cells('A1:F1')
ws['A1'].font = Font(size=16, bold=True)
ws['A1'].alignment = Alignment(horizontal='center')

# 表头
headers = ['镜头', '场景', '景别', 'Prompt', 'Negative Prompt', '采样参数']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col, value=header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    cell.font = Font(bold=True, color='FFFFFF')
    cell.alignment = Alignment(horizontal='center', vertical='center')

# 定义角色一致性特征
lin_yuan = "42-year-old Asian male captain, square jaw, short graying hair, deep eyes, 3cm scar on left eyebrow, dark blue captain uniform with four gold stripes, Earth badge on chest, portable data terminal on belt, antique mechanical watch on left wrist"

su_qing = "35-year-old Asian female scientist, shoulder-length black hair, sharp eyes, wearing anti-radiation glasses, white research uniform, quantum computer interface bracelet on left arm, holographic data analyzer in hand, calluses on fingers from data pen"

qiming_ai = "holographic projection of glowing geometric sphere, blue-white luminescence, neutral synthetic voice visualization"

space_station = "abandoned space station covered in unknown black substance, floating in deep space, mysterious mathematical symbols glowing on surface"

# 定义镜头数据
shots = [
    (1, "驾驶舱", "全景", 
     f"Deep space background, spaceship cockpit interior, cool blue-purple lighting, metallic surfaces, holographic displays glowing, {lin_yuan} standing at main control panel, back view, cinematic sci-fi atmosphere, 8k, highly detailed",
     "blurry, low quality, cartoon, anime, earth, planet surface"),
    
    (2, "驾驶舱", "特写", 
     "Close-up of holographic display screen, complex waveform patterns unfolding, perfect mathematical sequence visualization, prime number encryption patterns, blue-purple glowing data streams, futuristic UI design, 8k, sharp details",
     "blurry, distorted, low resolution, hand-drawn, sketch"),
    
    (3, "驾驶舱", "近景", 
     f"Side view of {lin_yuan}, staring at screen, scar on left eyebrow visible under blue light, eyes gradually lighting up with excitement, dramatic side lighting, cinematic depth of field, 8k, photorealistic",
     "front view, blurry face, cartoon style, low quality"),
    
    (4, "驾驶舱", "特写", 
     f"Extreme close-up of antique mechanical watch on wrist, ticking second hand, engraved text on dial, metallic texture, warm lighting contrasting with cool background, macro photography style, 8k, detailed",
     "digital watch, smart watch, blurry, low detail"),
    
    (5, "驾驶舱", "中景", 
     f"{lin_yuan} fingers typing on virtual holographic keyboard, pressing confirmation button, blue light reflecting on hands, futuristic interface, side angle, motion blur on fingers, 8k, cinematic",
     "blurry hands, cartoon, low quality, earth background"),
    
    (6, "实验室", "全景", 
     f"Science laboratory interior, {su_qing} standing at holographic data platform, white uniform under blue lighting, fingers sliding through data streams, high-tech equipment, glass walls, 8k, atmospheric",
     "messy, cluttered, low tech, blurry"),
    
    (7, "实验室", "特写", 
     f"Close-up of {su_qing} anti-radiation glasses, lenses reflecting scrolling code, sharp focused eyes behind glasses, blue light reflection, mysterious atmosphere, 8k, detailed",
     "sunglasses, blurry reflection, cartoon eyes"),
    
    (8, "实验室", "近景", 
     f"{su_qing} comparing two waveform patterns on holographic display, furrowed brow, professional concentration, white uniform, blue ambient light, side profile, 8k, photorealistic",
     "front view, smiling, casual clothes, low quality"),
    
    (9, "实验室", "快速剪辑", 
     "Abstract visualization of red code spreading like virus, consuming original navigation system code, digital corruption effect, red and blue contrast, cyberpunk aesthetic, glitch effects, 8k, dynamic",
     "static image, blurry, low contrast"),
    
    (10, "实验室", "特写", 
     f"{su_qing} face close-up, fear flashing in eyes, remembering mother's lab explosion, quick flashback overlay of fire and destruction, emotional distress, dramatic lighting, 8k, cinematic",
     "calm expression, smiling, low quality"),
    
    (11, "驾驶舱", "近景", 
     f"{lin_yuan} about to respond when cockpit lights suddenly turn red, alarm lighting, surprised expression, {qiming_ai} visible in background, tension building, 8k, dramatic",
     "calm lighting, smiling, earth visible"),
    
    (12, "驾驶舱", "特写", 
     "Close-up of control panel, all buttons turning gray, screen displaying 'AUTONAVIGATION ENGAGED' in red text, system lockdown, futuristic UI, warning symbols, 8k, detailed",
     "working buttons, green lights, low quality"),
    
    (13, "走廊", "跟随", 
     f"Overhead tracking shot, {lin_yuan} running through spaceship corridor, {su_qing} running from opposite direction, meeting in middle, red alarm lights flashing, metallic walls, urgent atmosphere, motion blur, 8k, dynamic",
     "static shot, empty corridor, earth gravity"),
    
    (14, "走廊", "双人近景", 
     f"{lin_yuan} and {su_qing} facing each other in corridor, red alarm lights pulsing, {su_qing} grabbing {lin_yuan} arm, intense confrontation, dramatic side lighting, 8k, cinematic",
     "smiling, casual clothes, earth background, low quality"),
    
    (15, "走廊", "特写", 
     f"{lin_yuan} flipping mechanical watch, revealing engraved text on back, 'LOOK BACK' inscription, metallic texture, red alarm light reflecting, emotional moment, macro shot, 8k, detailed",
     "digital watch, blurry text, cartoon"),
    
    (16, "引擎室", "全景", 
     "Wide shot of fusion engine room, massive reactor glowing with blue light, heat distortion in air, industrial sci-fi design, pipes and cables, atmospheric lighting, 8k, epic scale",
     "small room, earth technology, blurry"),
    
    (17, "引擎室", "近景", 
     f"{lin_yuan} wearing thermal gloves, grabbing red emergency control lever, {qiming_ai} hologram visible warning, blue engine glow illuminating face, tension, side angle, 8k, photorealistic",
     "no gloves, smiling, low quality"),
    
    (18, "引擎室", "特写", 
     f"{lin_yuan} hand trembling on control lever, red paint peeling off lever, thermal glove texture, blue engine light, emotional struggle visible, macro photography, 8k, detailed",
     "steady hand, no gloves, cartoon"),
    
    (19, "引擎室", "特写", 
     f"{lin_yuan} eyes close-up, hesitation turning to determination, reflection of blue engine light in pupils, emotional transformation, shallow depth of field, 8k, cinematic",
     "closed eyes, blurry, low quality"),
    
    (20, "引擎室", "特写", 
     "Control lever being pulled down, metal friction sparks flying, mechanical detail, industrial texture, dramatic lighting, action moment, macro shot, 8k, detailed",
     "static lever, no sparks, cartoon"),
    
    (21, "引擎室", "全景", 
     f"Engine room returning to normal blue-white lighting, {lin_yuan} kneeling on floor catching breath, relief on face, massive engine in background, atmospheric, 8k, cinematic",
     "red alarm still on, standing, smiling, low quality"),
    
    (22, "驾驶舱", "全景", 
     f"Slow push-in shot, {lin_yuan} and {su_qing} standing at main control panel, starfield flowing outside window, peaceful atmosphere, blue-white lighting, wide cinematic composition, 8k",
     "empty cockpit, earth visible, blurry"),
    
    (23, "驾驶舱", "特写", 
     "Screen displaying two options: A. Resume auto-navigation B. Manual control转向, futuristic UI design, blue glow, decision moment, 8k, detailed interface",
     "blank screen, error message, low quality"),
    
    (24, "驾驶舱", "近景", 
     f"{lin_yuan} side profile, looking at mechanical watch then at screen showing {space_station}, internal conflict, starlight reflection, emotional moment, 8k, cinematic",
     "smiling, casual, earth background"),
    
    (25, "驾驶舱", "特写", 
     f"{lin_yuan} finger hovering over button, slight trembling, anticipation, macro shot of finger and button, dramatic lighting, tension, 8k, detailed",
     "pressing button immediately, steady hand, cartoon"),
    
    (26, "驾驶舱", "特写", 
     "Slow motion: finger pressing button B, confirmation light activating, futuristic button design, blue glow, decisive moment, high speed photography effect, 8k, dramatic",
     "pressing wrong button, blurry, low quality"),
    
    (27, "飞船外部", "全景", 
     "Spaceship turning in deep space, starfield rotating outside window, smooth maneuver, blue engine thrusters firing, cinematic space shot, epic scale, 8k, majestic",
     "static ship, earth visible, blurry"),
    
    (28, "驾驶舱", "特写", 
     f"Screen showing distance numbers increasing: 12 light-hours, 15, 20... {space_station} getting smaller on display, departure visualization, 8k, detailed UI",
     "numbers decreasing, error, low quality"),
    
    (29, "驾驶舱", "近景", 
     f"{su_qing} smiling, first time showing vulnerability, relief and respect in eyes, soft lighting, emotional release, 8k, photorealistic portrait",
     "serious expression, frowning, low quality"),
    
    (30, "驾驶舱", "特写", 
     f"Mechanical watch close-up, second hand continuing to move, dial reflecting starlight, metallic gleam, symbolic moment, macro photography, 8k, detailed",
     "stopped watch, digital display, blurry"),
    
    (31, "深空", "远景", 
     "Wide shot of spaceship sailing into deep space, mysterious space station disappearing from view, vast starfield, infinite cosmos, philosophical ending, 8k, epic, majestic",
     "space station still visible, earth, blurry, low quality")
]

# 填充数据
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for row_idx, (shot_num, scene, shot_type, prompt, negative) in enumerate(shots, 4):
    ws.cell(row=row_idx, column=1, value=shot_num)
    ws.cell(row=row_idx, column=2, value=scene)
    ws.cell(row=row_idx, column=3, value=shot_type)
    ws.cell(row=row_idx, column=4, value=prompt)
    ws.cell(row=row_idx, column=5, value=negative)
    ws.cell(row=row_idx, column=6, value="Steps: 30, CFG: 7-8, Sampler: DPM++ 2M Karras, Size: 1024x576")
    
    # 设置样式
    for col in range(1, 7):
        cell = ws.cell(row=row_idx, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='top', wrap_text=True)

# 保存
wb.save('05_ComfyUI提示词.xlsx')
print("✅ ComfyUI提示词表已生成：05_ComfyUI提示词.xlsx")
print(f"📊 总计：{len(shots)} 个镜头")
