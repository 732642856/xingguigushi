#!/usr/bin/env python3
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# 分镜数据：(场景, 镜头编号, 景别, 机位, 运动, 时长, 画面内容, 声音/对白, 备注)
SHOTS = [
    # 第1场 菜园清晨
    (
        "1\n菜园清晨",
        "001",
        "远景→全景",
        "正面",
        "缓慢推进",
        8,
        "清晨菜园全景，薄雾笼罩，阳光从地平线升起，老周背着水壶走进画面开始浇水",
        "轻柔的鸟鸣，远处犬吠",
        "开场定调，宁静治愈",
    ),
    (
        "1\n菜园清晨",
        "002",
        "特写",
        "正面",
        "固定",
        3,
        "水壶倾斜浇水，番茄上的露珠晶莹，阳光穿过露珠折射彩虹色",
        "水滴声",
        "展现田园细节美",
    ),
    (
        "1\n菜园清晨",
        "003",
        "中景",
        "正面侧方",
        "手持跟拍",
        6,
        "老周浇灌茄子、黄瓜，认真照顾每株菜，额头上沁出汗水",
        "水流声",
        "展现劳作细节",
    ),
    # 第2场 邻居小孩
    (
        "2\n邻居小孩",
        "004",
        "近景",
        "篱笆外",
        "固定",
        4,
        "小白趴在篱笆上，大眼睛好奇往里看，手里拿着狗尾草",
        "孩童轻笑",
        "建立角色",
    ),
    (
        "2\n邻居小孩",
        "005",
        "中景",
        "菜园内",
        "固定",
        5,
        "老周抬头看到小白，微笑抬手邀请：进来看看吧",
        "老周：小朋友，想进来看看吗？",
        "温暖互动",
    ),
    (
        "2\n邻居小孩",
        "006",
        "特写",
        "正面",
        "固定",
        3,
        "小白的小手触碰红色的番茄，眼睛发亮",
        "哇！",
        "童年好奇",
    ),
    (
        "2\n邻居小孩",
        "007",
        "中景",
        "菜园内",
        "缓慢推进",
        6,
        "老周牵着小手介绍：这是番茄，这是茄子...",
        "老周：慢点，别碰到花",
        "传承时刻",
    ),
    # 第3场 共进午餐
    (
        "3\n厨房",
        "008",
        "全景",
        "厨房门口",
        "固定",
        5,
        "厨房全景，桌上摆着炒土豆丝、番茄鸡蛋汤，窗外阳光温暖",
        "锅碗声",
        "家庭温暖",
    ),
    (
        "3\n厨房",
        "009",
        "中景",
        "桌旁",
        "固定",
        8,
        "两人坐对面吃饭，小白狼吞虎咽，老周夹菜给他",
        "老周：多吃点。小白：爷爷做的饭真好吃！",
        "温馨用餐",
    ),
    (
        "3\n厨房",
        "010",
        "特写",
        "正面",
        "固定",
        4,
        "小白夹菜送入口中，米饭沾在脸上，幸福的笑容",
        "嗯~",
        "可爱时刻",
    ),
    (
        "3\n厨房",
        "011",
        "近景",
        "正面",
        "固定",
        6,
        "老周讲述往事，眼神看向远方：那是三十年前...",
        "老周：我以前教语文...",
        "回忆引入",
    ),
    # 第4场 树下交谈
    (
        "4\n院子",
        "012",
        "全景",
        "老槐树",
        "固定",
        6,
        "老槐树下斑驳光影，摇椅和小凳，老周和小白",
        "蝉鸣",
        "午后时光",
    ),
    (
        "4\n院子",
        "013",
        "中景",
        "侧面",
        "固定",
        8,
        "小白托腮听得入神，眼睛发亮：然后呢？",
        "小白：后来呢后来呢？",
        "童真",
    ),
    (
        "4\n院子",
        "014",
        "近景",
        "正面",
        "固定",
        6,
        "老周微笑讲述，皱纹里藏着故事：那个学生现在是大医生了",
        "老周：是啊，真好",
        "满足",
    ),
    # 第5场 告别
    (
        "5\n告别",
        "015",
        "远景",
        "村口",
        "固定",
        5,
        "夕阳下村口，小白背着书包站着，金色光芒笼罩",
        "归巢鸟鸣",
        "黄昏诗意",
    ),
    (
        "5\n告别",
        "016",
        "中景",
        "正面",
        "固定",
        4,
        "小白挥手：爷爷再见！",
        "小白：爷爷明天见！",
        "告别",
    ),
    (
        "5\n告别",
        "017",
        "近景",
        "正面",
        "固定",
        4,
        "老周目送，微笑挥手，眼中泛光",
        "老周：诶，慢点走",
        "不舍",
    ),
    (
        "5\n菜园",
        "018",
        "远景→全景",
        "菜园",
        "缓慢拉远",
        8,
        "老周走回菜园，夕阳把影子拉得很长，开始浇水",
        "水流声，远处犬吠",
        "治愈结尾",
    ),
]

# 创建Excel
wb = Workbook()
ws = wb.active
ws.title = "分镜头脚本"

# 标题
ws["A1"] = "《田园风光》分镜头脚本"
ws.merge_cells("A1:I1")
ws["A1"].font = Font(size=16, bold=True)
ws["A1"].alignment = Alignment(horizontal="center")
ws.row_dimensions[1].height = 25

# 副标题
ws["A2"] = "时长：5分钟 | 18个镜头"
ws.merge_cells("A2:I2")
ws["A2"].font = Font(size=10, italic=True)
ws["A2"].alignment = Alignment(horizontal="center")

# 表头
headers = [
    "场景",
    "镜头编号",
    "景别",
    "机位",
    "运动",
    "时长(秒)",
    "画面内容",
    "声音/对白",
    "备注",
]
header_fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# 边框
border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# 场景颜色映射
scene_colors = {
    "1": "FFF3E0",  # 橙色
    "2": "E8F5E9",  # 绿色
    "3": "FFEBEE",  # 粉色
    "4": "E3F2FD",  # 蓝色
    "5": "F3E5F5",  # 紫色
}

# 数据
for row_idx, row_data in enumerate(SHOTS, 4):
    scene_num = str(row_data[0]).split("\n")[0]
    for col in range(1, 10):
        cell = ws.cell(row=row_idx, column=col, value=row_data[col - 1])
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.fill = PatternFill(
            start_color=scene_colors.get(scene_num, "FFFFFF"),
            end_color=scene_colors.get(scene_num, "FFFFFF"),
            fill_type="solid",
        )
    ws.row_dimensions[row_idx].height = 60

# 列宽
ws.column_dimensions["A"].width = 12
ws.column_dimensions["B"].width = 8
ws.column_dimensions["C"].width = 10
ws.column_dimensions["D"].width = 10
ws.column_dimensions["E"].width = 10
ws.column_dimensions["F"].width = 8
ws.column_dimensions["G"].width = 40
ws.column_dimensions["H"].width = 30
ws.column_dimensions["I"].width = 25

# 保存
output = "/Volumes/ssd4t/ai/openwork/cv01/剧本项目/田园风光_20260412/04_分镜.xlsx"
wb.save(output)
print(f"✅ 已生成：{output}")
print(f"📊 共 {len(SHOTS)} 个镜头")
