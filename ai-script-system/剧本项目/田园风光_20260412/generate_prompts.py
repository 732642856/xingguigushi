#!/usr/bin/env python3
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# 角色一致性特征
LAO_ZHOU = "65-year-old Chinese grandfather, gray-white hair, kind weathered face with gentle wrinkles, wearing loose white shirt and dark pants, warm smile, carrying watering can"

XIAO_BAI = "8-year-old Chinese boy, round face with big curious eyes, fair skin, wearing cartoon T-shirt and shorts, playful expression, small hands reaching out"

VEGETABLES = "lush vegetable garden with tomatoes, cucumbers, eggplants, morning dew on leaves, warm golden sunrise lighting"

# 镜头定义
SHOTS = [
    (
        1,
        "1",
        "菜园清晨",
        "远景→全景",
        None,
        f"Wide establishing shot of rural vegetable garden at dawn, misty atmosphere, warm golden sunrise breaking through, {VEGETABLES}, serene farming landscape, cinematic composition, photorealistic, highly detailed, 8k",
    ),
    (
        2,
        "1",
        "菜园清晨",
        "特写",
        None,
        f"Extreme close-up of watering can pouring water on tomato, morning dew glistening on ripe red tomatoes, light refracting through water drops, rainbow effect, shallow depth of field, macro photography, photorealistic, highly detailed, 8k",
    ),
    (
        3,
        "1",
        "菜园清晨",
        "中景",
        "老周",
        f"Medium shot of {LAO_ZHOU} carefully watering eggplants and cucumbers, gentle focused expression, morning light casting soft shadows, peaceful farming routine, cinematic composition, photorealistic, highly detailed, 8k",
    ),
    (
        4,
        "2",
        "邻居小孩",
        "近景",
        "小白",
        f"Close-up of {XIAO_BAI} peeking over wooden fence, big curious eyes looking at vegetable garden, eager expression, childhood wonder, morning light from behind, shallow depth of field, photorealistic, highly detailed, 8k",
    ),
    (
        5,
        "2",
        "邻居小孩",
        "中景",
        "老周+小白",
        f"Medium shot of {LAO_ZHOU} looking up at fence, inviting gesture, {XIAO_BAI} on other side, warm interaction between generations, soft morning lighting, cinematic two-shot, photorealistic, highly detailed, 8k",
    ),
    (
        6,
        "2",
        "邻居小孩",
        "特写",
        "小白",
        f"Extreme close-up of {XIAO_BAI} small hand touching red tomato, wonder and delight in eyes, gentle touch, morning light illuminating curious face, macro photography, photorealistic, highly detailed, 8k",
    ),
    (
        7,
        "2",
        "菜园",
        "中景",
        "老周+小白",
        f"Medium shot of {LAO_ZHOU} holding {XIAO_BAI}'s hand, pointing at vegetables, {VEGETABLES} in background, teaching moment, warm afternoon light, cinematic composition, photorealistic, highly detailed, 8k",
    ),
    (
        8,
        "3",
        "厨房",
        "全景",
        None,
        f"Wide shot of rustic country kitchen, wooden table with home-cooked dishes: stir-fried potatoes, tomato egg soup, warm yellow lighting from window, cozy domestic atmosphere, photorealistic, highly detailed, 8k",
    ),
    (
        9,
        "3",
        "厨房",
        "中景",
        "老周+小白",
        f"Medium shot of {LAO_ZHOU} and {XIAO_BAI} eating together at table, warm conversation, home cooking, soft interior lighting, generational bond, cinematic two-shot, photorealistic, highly detailed, 8k",
    ),
    (
        10,
        "3",
        "厨房",
        "特写",
        "小白",
        f"Extreme close-up of {XIAO_BAI} eating enthusiastically, rice on cheeks, delighted expression, food in mouth, humorous but endearing, warm lighting, shallow depth of field, macro shot, photorealistic, highly detailed, 8k",
    ),
    (
        11,
        "3",
        "厨房",
        "近景",
        "老周",
        f"Close-up of {LAO_ZHOU} with nostalgic expression, eyes looking upward remembering past, warm ambient lighting, storytelling mood, cinematic close-up, photorealistic, highly detailed, 8k",
    ),
    (
        12,
        "4",
        "院子",
        "全景",
        None,
        f"Wide shot of old sycamore tree in courtyard, dappled sunlight through leaves, wooden chairs, peaceful afternoon, {LAO_ZHOU} and {XIAO_BAI} in scene, warm golden hour lighting, photorealistic, highly detailed, 8k",
    ),
    (
        13,
        "4",
        "院子",
        "中景",
        "小白",
        f"Medium shot of {XIAO_BAI} sitting with chin on hands, listening attentively, big eyes shining with curiosity, warm afternoon light on face, generational storytelling, cinematic composition, photorealistic, highly detailed, 8k",
    ),
    (
        14,
        "4",
        "院子",
        "近景",
        "老周",
        f"Close-up of {LAO_ZHOU} smiling warmly while telling stories, gentle expression, wrinkles showing warmth, afternoon light creating soft shadows, emotional moment, photorealistic, highly detailed, 8k",
    ),
    (
        15,
        "5",
        "告别",
        "远景",
        None,
        f"Wide establishing shot of village entrance at sunset, warm orange-gold sky, silhouette of small figure with backpack, serene rural landscape, golden hour magic hour lighting, cinematic composition, photorealistic, highly detailed, 8k",
    ),
    (
        16,
        "5",
        "告别",
        "中景",
        "小白",
        f"Medium shot of {XIAO_BAI} waving goodbye, backpack on shoulders, joyful expression, turning to walk away, warm sunset glow, cinematic medium shot, photorealistic, highly detailed, 8k",
    ),
    (
        17,
        "5",
        "告别",
        "近景",
        "老周",
        f"Close-up of {LAO_ZHOU} watching, gentle smile, slight melancholy, warm golden light illuminating face, heartfelt goodbye, cinematic portrait, photorealistic, highly detailed, 8k",
    ),
    (
        18,
        "5",
        "菜园",
        "远景→全景",
        "老周",
        f"Wide shot of {LAO_ZHOU} walking back to vegetable garden at sunset, long shadow on ground, watering can in hand, warm sunset casting golden light, peaceful ending, cinematic wide composition, photorealistic, highly detailed, 8k",
    ),
]

# 负面提示词
NEGATIVE = "cartoon, anime, illustration, 3d render, cgi, deformed, blurry, low quality, duplicate, watermark, signature, text, cropped, worst quality, low resolution, disfigured, bad anatomy"

# 参数
PARAMETERS = "Model: Animagine XL 3.1 | Steps: 28 | CFG: 7 | Sampler: DPM++ 2M Karras | Resolution: 1024x576"

# 创建Excel
wb = Workbook()
ws = wb.active
ws.title = "ComfyUI提示词"

# 标题
ws["A1"] = "《田园风光》ComfyUI提示词表"
ws.merge_cells("A1:F1")
ws["A1"].font = Font(size=16, bold=True)
ws["A1"].alignment = Alignment(horizontal="center")

# 表头
headers = ["镜头", "场次", "场景", "景别", "Prompt", "Negative Prompt", "采样参数"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

# 边框
border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# 数据
for row_idx, (shot, scene, location, shot_type, subject, prompt) in enumerate(SHOTS, 4):
    ws.cell(row=row_idx, column=1, value=shot)
    ws.cell(row=row_idx, column=2, value=scene)
    ws.cell(row=row_idx, column=3, value=location)
    ws.cell(row=row_idx, column=4, value=shot_type)
    ws.cell(row=row_idx, column=5, value=prompt)
    ws.cell(row=row_idx, column=6, value=NEGATIVE)
    ws.cell(row=row_idx, column=7, value=PARAMETERS)

    for col in range(1, 8):
        cell = ws.cell(row=row_idx, column=col)
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.row_dimensions[row_idx].height = 100

# 列宽
ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 6
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 12
ws.column_dimensions["E"].width = 80
ws.column_dimensions["F"].width = 40
ws.column_dimensions["G"].width = 35

# 保存
output = (
    "/Volumes/ssd4t/ai/openwork/cv01/剧本项目/田园风光_20260412/05_ComfyUI提示词.xlsx"
)
wb.save(output)
print(f"✅ 已生成：{output}")
print(f"📊 共 {len(SHOTS)} 个镜头")
